# Parse the scraped Election Night Reporting exports (see
# scrape_enr_exports.py) into county- and precinct-level CSVs for
# elections 2020-2024, following the same conventions as
# parse_2026_primary.py.
#
# Inputs per election (in ../openelections-sources-nd/nd_enr_exports/<tag>/):
#   - manifest.json : contest_id -> {title, party} (the workbooks don't
#     carry the party)
#   - precinct/*.xlsx : one workbook per contest, one sheet per county;
#     each sheet has a candidate header row and one row per precinct,
#     followed by a TOTALS row with the county total
#   - all_contests_county.csv : statewide-only totals, used as a cross-check
#
# County-level rows come from the workbooks' TOTALS rows, so county totals
# equal the sum of their precincts by construction.
import csv
import glob
import json
import os
import re

import openpyxl

SRC_ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        '../openelections-sources-nd/nd_enr_exports')
REPO = os.path.dirname(os.path.abspath(__file__))

COUNTY_FIXUPS = {
    'Mchenry': 'McHenry',
    'Mcintosh': 'McIntosh',
    'Mckenzie': 'McKenzie',
    'Mclean': 'McLean',
}

# Party codes as they appear in the exports -> repo convention
PARTY_FIXUPS = {
    'DNPL': 'DEM',
    'DFL': 'DEM',
    'NON': '',
}

# Office names standardized to match the existing files.
OFFICE_RENAMES = {
    'State Representative': 'State House',
    'State Senator': 'State Senate',
    'United States Senator': 'U.S. Senate',
    'Representative in Congress': 'U.S. House',
    'President & Vice-President of the United States': 'President',
}

PARTY_SUFFIX = re.compile(r' (Republican|Democratic-NPL|Democratic|Libertarian)$')
NONPARTISAN_SUFFIX = re.compile(r' Nonpartisan$')

# Elections whose archived workbooks are an incomplete snapshot: the 2006
# general's U.S. House TOTALS sum to 120,105 vs a certified 142,934, and
# several counties' precinct detail is missing outright. For these, county
# rows come from the page-displayed totals harvested by harvest_enr_counts.py
# (which there match the site CSV/certified results). Elsewhere the workbook
# TOTALS are authoritative (2004 matches the site CSV exactly and sits within
# 18 votes of the certified canvass; the 2008 general's page totals reconcile
# with its workbooks identically). Contests with no workbook at all always
# fall back to page votes.
PAGE_VOTE_ELECTIONS = {'20061107__general'}


def clean_name(name):
    return re.sub(r'\s+', ' ', str(name).strip())


def clean_county(name):
    name = name.strip()
    if name.endswith(' County'):
        name = name[:-len(' County')]
    return COUNTY_FIXUPS.get(name, name)


def parse_contest(contest_name, party):
    """Return (office, district, party) for a contest name."""
    party = PARTY_FIXUPS.get(party, party)
    party = party if party in ('DEM', 'REP') else ''
    office = NONPARTISAN_SUFFIX.sub('', PARTY_SUFFIX.sub('', clean_name(contest_name)))
    district = ''
    m = re.search(r'\bState (Senator|Representative)(.*?) District (\d+[a-zA-Z]?)$', office)
    if m:
        office = 'State ' + ('Senate' if m.group(1) == 'Senator' else 'House') + m.group(2)
        # leading zeros stripped, letter suffix kept: "04a" -> "4a"
        district = re.sub(r'^0+(?=.)', '', m.group(3))
    office = OFFICE_RENAMES.get(office, office)
    if office == 'U.S. House':
        district = '1'
    return office.strip(), district, party


def parse_precinct_xlsx(xlsx_path, manifest, use_page_votes=False):
    """Return (precinct_rows, county_rows) for one contest's workbook.

    When use_page_votes is set (elections in PAGE_VOTE_ELECTIONS, where the
    archived workbooks are an incomplete snapshot) and the manifest carries
    'county_page_votes' harvested by harvest_enr_counts.py, those page
    totals replace the workbook TOTALS as the county-level rows."""
    contest_id = os.path.splitext(os.path.basename(xlsx_path))[0]
    party = manifest[contest_id]['party']
    # general elections: parties appear only in the page HTML, harvested
    # per candidate by harvest_enr_party.py
    candidate_parties = manifest[contest_id].get('candidate_parties', {})
    page_votes = manifest[contest_id].get('county_page_votes', {}) if use_page_votes else {}
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    precinct_rows, county_rows = [], []
    for sheet in wb.sheetnames:
        county = clean_county(sheet)
        # the workbook header cell is just the bare office ("State
        # Representative") — the manifest title carries the district
        office, district, _ = parse_contest(manifest[contest_id]['title'], party)
        candidates = None
        candidates = None
        for row in wb[sheet].iter_rows(values_only=True):
            cells = [c if c is not None else '' for c in row]
            # contest header rows have 'Precinct' in the second column
            if str(cells[1]).strip() == 'Precinct':
                candidates = [clean_name(c) for c in cells[2:]]
                continue
            if candidates is None or not str(cells[1]).strip():
                continue
            precinct = clean_name(str(cells[1]))
            if precinct.startswith('Combined Precinct'):
                # 2000s-era exports are county-level: this row duplicates
                # the TOTALS row, there is no precinct detail
                continue
            if precinct == 'TOTALS':
                if county in page_votes:
                    # page-displayed county totals replace the workbook's
                    # (incomplete) TOTALS row
                    for candidate, votes in page_votes[county].items():
                        county_rows.append((county, office, district,
                                            candidate_parties.get(candidate, party), candidate,
                                            int(votes)))
                else:
                    for candidate, votes in zip(candidates, cells[2:]):
                        county_rows.append((county, office, district,
                                            candidate_parties.get(candidate, party), candidate,
                                            0 if str(votes).strip() in ('', 'None') else int(votes)))
            else:
                for candidate, votes in zip(candidates, cells[2:]):
                    precinct_rows.append((county, precinct, office, district,
                                          candidate_parties.get(candidate, party), candidate,
                                          0 if str(votes).strip() in ('', 'None') else int(votes)))
    wb.close()
    return precinct_rows, county_rows


def parse_page_only(contest_id, manifest):
    """County-level rows for a contest that has page-displayed totals but
    no workbook export (e.g. judgeships on some platforms)."""
    party = manifest[contest_id]['party']
    candidate_parties = manifest[contest_id].get('candidate_parties', {})
    office, district, _ = parse_contest(manifest[contest_id]['title'], party)
    rows = []
    for county, votes in manifest[contest_id].get('county_page_votes', {}).items():
        for candidate, v in votes.items():
            rows.append((county, office, district,
                         candidate_parties.get(candidate, party), candidate, int(v)))
    return rows


def aggregate_and_write(rows, path, header):
    totals = {}
    for row in rows:
        totals[row[:-1]] = totals.get(row[:-1], 0) + row[-1]
    with open(path, 'wt', newline='') as f:
        w = csv.writer(f)
        w.writerow(header)
        for key in sorted(totals):
            w.writerow([*key, totals[key]])
    print(f'wrote {len(totals)} rows to {path}')
    return totals


def main(elections=None, write_precinct=True):
    """Parse scraped exports into CSVs. write_precinct=False keeps the
    precinct data out of the repo (used when the repo already has a
    precinct file for that election from another source)."""
    for election_dir in sorted(glob.glob(os.path.join(SRC_ROOT, '*'))):
        tag = os.path.basename(election_dir)
        if elections and tag not in elections:
            continue
        manifest_path = os.path.join(election_dir, 'manifest.json')
        if not os.path.exists(manifest_path):
            print(f'{tag}: no manifest yet, skipping')
            continue
        with open(manifest_path) as f:
            manifest = json.load(f)
        year, date, kind = tag[:4], tag.split('__')[0], tag[10:]
        os.makedirs(f'{REPO}/{year}', exist_ok=True)
        precinct_rows, county_rows = [], []
        paths = sorted(glob.glob(os.path.join(election_dir, 'precinct', '*.xlsx')))
        missing = set(manifest) - {os.path.splitext(os.path.basename(p))[0] for p in paths}
        if missing:
            print(f'{tag}: {len(missing)} contests have no export: {sorted(missing)[:10]}')
        for path in paths:
            p, c = parse_precinct_xlsx(path, manifest,
                                       use_page_votes=tag in PAGE_VOTE_ELECTIONS)
            precinct_rows.extend(p)
            county_rows.extend(c)
        # contests with no workbook but page-displayed totals still
        # contribute county-level rows
        for contest_id in sorted(set(manifest) - {os.path.splitext(os.path.basename(p))[0]
                                                  for p in paths}):
            if manifest[contest_id].get('county_page_votes'):
                county_rows.extend(parse_page_only(contest_id, manifest))
        county_totals = aggregate_and_write(county_rows,
            f'{REPO}/{year}/{date}__nd__{kind}__county.csv',
            ['county', 'office', 'district', 'party', 'candidate', 'votes'])
        if not precinct_rows:
            # county-level-only platform (2000s): no precinct file to write
            print(f'{tag}: exports carry no precinct detail, no precinct CSV')
            precinct_totals = {}
        else:
            precinct_totals = aggregate_and_write(precinct_rows,
                (f'{REPO}/{year}/{date}__nd__{kind}__precinct.csv' if write_precinct
                 else os.path.join(election_dir, 'precinct.csv')),
                ['county', 'precinct', 'office', 'district', 'party', 'candidate', 'votes'])

        # cross-check 1: precinct sums must equal county TOTALS (keys only in
        # one file are reported separately: contests with page-derived county
        # rows but no workbook carry no precinct detail)
        if precinct_totals:
            prec_by_county = {}
            for key, votes in precinct_totals.items():
                k = (key[0], *key[2:])
                prec_by_county[k] = prec_by_county.get(k, 0) + votes
            common = set(county_totals) & set(prec_by_county)
            mism = sum(1 for k in common if county_totals[k] != prec_by_county[k])
            only_prec = len(set(prec_by_county) - common)
            only_county = len(set(county_totals) - common)
            print(f'{tag}: county/precinct mismatches: {mism} '
                  f'(county-only keys: {only_county}, precinct-only keys: {only_prec})')

        # cross-check 2: county TOTALS for statewide contests must equal the
        # site's statewide CSV
        statewide_src = os.path.join(election_dir, 'all_contests_county.csv')
        if os.path.exists(statewide_src):
            state = {}
            with open(statewide_src, newline='', encoding='utf-8-sig') as f:
                for r in csv.DictReader(f):
                    if not r['ContestName'].strip():
                        continue
                    office, district, party = parse_contest(r['ContestName'], r['PartyCode'])
                    # legislative districts are in AreaNum ("District 10"),
                    # not in the contest name
                    m = re.search(r'District (\d+[a-zA-Z]?)', r.get('AreaNum', ''))
                    if m and not district:
                        district = re.sub(r'^0+(?=.)', '', m.group(1))
                    key = (office, district, party, clean_name(r['CandidateName']))
                    state[key] = state.get(key, 0) + (0 if r['CandidateVotes'].strip() == 'undefined'
                                                      else int(r['CandidateVotes']))
            bad = checked = 0
            for (office, district, party, candidate), v in state.items():
                s = sum(votes for key, votes in county_totals.items()
                        if key[1:] == (office, district, party, candidate))
                if s == 0:
                    continue  # county-specific contest names don't match
                checked += 1
                if s != v:
                    bad += 1
                    if bad <= 3:
                        print(f'  statewide mismatch {(office, district, party, candidate)}: '
                              f'site {v} vs parsed {s}')
            print(f'{tag}: statewide cross-check: {bad} mismatches / {checked} checked')


if __name__ == '__main__':
    main()